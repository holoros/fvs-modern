#!/usr/bin/env python3
"""
Generate an Excel template for FVS variant coefficient calibration.

This script creates a structured Excel workbook that guides researchers
through the process of specifying all coefficients needed for a new FVS
variant or recalibrating an existing one.

Usage:
    python generate_coefficient_template.py [output_file.xlsx]

Default output: variant_coefficients_template.xlsx
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_header_format():
    """Create formatting for header rows."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    return header_fill, header_font, header_alignment, border


def create_data_format():
    """Create formatting for data cells."""
    data_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    data_font = Font(size=11)
    data_alignment = Alignment(horizontal="right", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    return data_fill, data_font, data_alignment, border


def format_header_row(ws, row, columns):
    """Format a header row in the worksheet."""
    header_fill, header_font, header_alignment, border = create_header_format()

    for col_num, col_title in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = col_title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border


def create_species_list_sheet(wb):
    """Create the Species_List sheet."""
    ws = wb.create_sheet("Species_List", 0)

    columns = [
        "Array Index",
        "FIA Code",
        "Common Name",
        "Scientific Name",
        "Genus",
        "Max DBH (in)",
        "Max Height (ft)",
        "Longevity (yr)",
        "Shade Tolerance (0-1)"
    ]

    format_header_row(ws, 1, columns)

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 18

    # Add placeholder rows for first 10 species (user fills in the rest)
    data_fill, data_font, data_alignment, border = create_data_format()
    for i in range(1, 11):
        for col_num in range(1, len(columns) + 1):
            cell = ws.cell(row=i+1, column=col_num)
            cell.fill = data_fill
            cell.border = border
            if col_num == 1:  # Array Index
                cell.value = i

    # Add instructions in column below
    ws.cell(row=15, column=1).value = "Instructions:"
    ws.cell(row=16, column=1).value = "1. Fill in FIA species codes for each array index (1-108)"
    ws.cell(row=17, column=1).value = "2. Provide common and scientific names from USDA Forest Service"
    ws.cell(row=18, column=1).value = "3. Max DBH and Height from species literature or expert opinion"
    ws.cell(row=19, column=1).value = "4. Shade tolerance: 0=intolerant, 1=very tolerant"


def create_crown_coefficients_sheet(wb):
    """Create the Crown_Coefficients sheet."""
    ws = wb.create_sheet("Crown_Coefficients", 1)

    columns = [
        "Species Index",
        "Species Code",
        "Common Name",
        "BCR1 (Intercept)",
        "BCR2 (BA Sensitivity)",
        "BCR3 (Asymptote)",
        "BCR4 (Shape)"
    ]

    format_header_row(ws, 1, columns)

    # Set column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 25
    for col in 'DEFG':
        ws.column_dimensions[col].width = 18

    # Add placeholder rows
    data_fill, data_font, data_alignment, border = create_data_format()
    for i in range(1, 11):
        for col_num in range(1, len(columns) + 1):
            cell = ws.cell(row=i+1, column=col_num)
            cell.fill = data_fill
            cell.border = border
            if col_num == 1:  # Species Index
                cell.value = i

    # Add explanation
    ws.cell(row=15, column=1).value = "Crown Ratio Equation:"
    ws.cell(row=16, column=1).value = "CR = 10 * (BCR1/(1+BCR2*BA) + BCR3*(1-exp(BCR4*DBH)))"
    ws.cell(row=17, column=1).value = ""
    ws.cell(row=18, column=1).value = "Calibration Sources:"
    ws.cell(row=19, column=1).value = "- FIA plot data with crown ratio measurements"
    ws.cell(row=20, column=1).value = "- Regional forest inventory cruise data"
    ws.cell(row=21, column=1).value = "- Non-linear regression (nlme or lme4 in R)"


def create_dg_coefficients_sheet(wb):
    """Create the DG_Coefficients sheet."""
    ws = wb.create_sheet("DG_Coefficients", 2)

    columns = [
        "Species Index",
        "Species Code",
        "Common Name",
        "B1 (SI Sensitivity)",
        "B2 (DBH Sensitivity)",
        "COR (Bias Correction)"
    ]

    format_header_row(ws, 1, columns)

    # Set column widths
    for col in 'ABC':
        ws.column_dimensions[col].width = 18
    for col in 'DEF':
        ws.column_dimensions[col].width = 18

    # Add placeholder rows
    data_fill, data_font, data_alignment, border = create_data_format()
    for i in range(1, 11):
        for col_num in range(1, len(columns) + 1):
            cell = ws.cell(row=i+1, column=col_num)
            cell.fill = data_fill
            cell.border = border
            if col_num == 1:
                cell.value = i

    # Add explanation
    ws.cell(row=15, column=1).value = "Diameter Growth Equation:"
    ws.cell(row=16, column=1).value = "POTBAG = B1 * SITEAR * (1 - exp(-B2 * DBH))"
    ws.cell(row=17, column=1).value = "DG_predicted = DG_model * COR"
    ws.cell(row=18, column=1).value = ""
    ws.cell(row=19, column=1).value = "B1 typical range: 0.0005 - 0.0015"
    ws.cell(row=20, column=1).value = "B2 typical range: 0.03 - 0.15"
    ws.cell(row=21, column=1).value = "COR typical range: 0.8 - 1.2 (1.0 = no bias)"


def create_mortality_coefficients_sheet(wb):
    """Create the Mortality_Coefficients sheet."""
    ws = wb.create_sheet("Mortality_Coefficients", 3)

    columns = [
        "Species Index",
        "Species Code",
        "Common Name",
        "VARADJ (Shade Tolerance)",
        "MORT_B0",
        "MORT_B1",
        "MORT_B2"
    ]

    format_header_row(ws, 1, columns)

    # Set column widths
    for col in 'ABC':
        ws.column_dimensions[col].width = 18
    for col in 'DEFG':
        ws.column_dimensions[col].width = 18

    # Add placeholder rows
    data_fill, data_font, data_alignment, border = create_data_format()
    for i in range(1, 11):
        for col_num in range(1, len(columns) + 1):
            cell = ws.cell(row=i+1, column=col_num)
            cell.fill = data_fill
            cell.border = border
            if col_num == 1:
                cell.value = i

    # Add explanation
    ws.cell(row=15, column=1).value = "Mortality Equation:"
    ws.cell(row=16, column=1).value = "Annual Mortality = 1 / (1 + exp(-(MORT_B0 + MORT_B1*ln(DBH))))"
    ws.cell(row=17, column=1).value = "Survival = Annual_Survival ^ years"
    ws.cell(row=18, column=1).value = ""
    ws.cell(row=19, column=1).value = "VARADJ: 0=intolerant (high mortality in shade)"
    ws.cell(row=20, column=1).value = "         1=tolerant (low mortality in shade)"


def create_htdbh_coefficients_sheet(wb):
    """Create the HT_DBH_Coefficients sheet."""
    ws = wb.create_sheet("HT_DBH_Coefficients", 4)

    columns = [
        "Species Index",
        "Species Code",
        "Common Name",
        "b0 (Intercept)",
        "b1 (Scaling)",
        "b2 (Exponent)"
    ]

    format_header_row(ws, 1, columns)

    # Set column widths
    for col in 'ABC':
        ws.column_dimensions[col].width = 18
    for col in 'DEF':
        ws.column_dimensions[col].width = 18

    # Add placeholder rows
    data_fill, data_font, data_alignment, border = create_data_format()
    for i in range(1, 11):
        for col_num in range(1, len(columns) + 1):
            cell = ws.cell(row=i+1, column=col_num)
            cell.fill = data_fill
            cell.border = border
            if col_num == 1:
                cell.value = i

    # Add explanation
    ws.cell(row=15, column=1).value = "Curtis-Arney Equation:"
    ws.cell(row=16, column=1).value = "H = 4.5 + exp(b0 + b1/DBH^b2)"
    ws.cell(row=17, column=1).value = ""
    ws.cell(row=18, column=1).value = "Fit to: H vs DBH from FIA inventory"
    ws.cell(row=19, column=1).value = "Typical b0: -1.5 to -3.0"
    ws.cell(row=20, column=1).value = "Typical b1: 10 to 20"
    ws.cell(row=21, column=1).value = "Typical b2: 0.9 to 1.2"


def create_site_index_sheet(wb):
    """Create the Site_Index sheet."""
    ws = wb.create_sheet("Site_Index", 5)

    columns = [
        "Species Index",
        "Species Code",
        "Common Name",
        "Default SI (ft@50yr)",
        "SDI Maximum",
        "SI Curve Notes"
    ]

    format_header_row(ws, 1, columns)

    # Set column widths
    for col in 'ABC':
        ws.column_dimensions[col].width = 18
    for col in 'DEF':
        ws.column_dimensions[col].width = 18

    # Add placeholder rows
    data_fill, data_font, data_alignment, border = create_data_format()
    for i in range(1, 11):
        for col_num in range(1, len(columns) + 1):
            cell = ws.cell(row=i+1, column=col_num)
            cell.fill = data_fill
            cell.border = border
            if col_num == 1:
                cell.value = i

    # Add explanation
    ws.cell(row=15, column=1).value = "Site Index (SI): Height at base age 50 years"
    ws.cell(row=16, column=1).value = "Higher SI = more productive site"
    ws.cell(row=17, column=1).value = ""
    ws.cell(row=18, column=1).value = "Default SI used when user doesn't specify"
    ws.cell(row=19, column=1).value = "SDI (Stand Density Index) Max: self-thinning threshold"


def create_instructions_sheet(wb):
    """Create the Instructions sheet."""
    ws = wb.create_sheet("Instructions", 6)
    ws.column_dimensions['A'].width = 100

    instructions = [
        ("FVS VARIANT COEFFICIENT CALIBRATION GUIDE", "Title"),
        ("", ""),
        ("OVERVIEW", "Section"),
        ("This workbook guides you through calibrating a new FVS variant or recalibrating",
         "Normal"),
        ("an existing variant. FVS requires species-specific coefficients for:", "Normal"),
        ("  1. Crown ratio models", "Normal"),
        ("  2. Diameter growth models", "Normal"),
        ("  3. Mortality/survival models", "Normal"),
        ("  4. Height-diameter relationships", "Normal"),
        ("  5. Site index curves", "Normal"),
        ("", ""),
        ("DATA REQUIREMENTS", "Section"),
        ("You will need:", "Normal"),
        ("  * FIA plot data with multiple measurements per tree over time (preferred)",
         "Normal"),
        ("  * Inventory plot clusters with diameter and height measurements", "Normal"),
        ("  * Stand history information (age, density, disturbance history)", "Normal"),
        ("  * Site variables (elevation, aspect, soil type)", "Normal"),
        ("", ""),
        ("WORKFLOW", "Section"),
        ("1. Complete the Species_List sheet", "Normal"),
        ("   - Map FIA species codes to array indices (1-108)", "Normal"),
        ("   - Provide common and scientific names", "Normal"),
        ("", ""),
        ("2. Fit the Crown Ratio Model", "Normal"),
        ("   - Use FIA data or cruise data with crown measurements", "Normal"),
        ("   - Non-linear regression: ln(CR) ~ BA + DBH", "Normal"),
        ("   - See R examples below for estimation code", "Normal"),
        ("", ""),
        ("3. Fit the Diameter Growth Model", "Normal"),
        ("   - Use repeated measures data (same trees over time)", "Normal"),
        ("   - Model: ln(DDS) ~ ln(DBH) + BA + SI", "Normal"),
        ("   - Use mixed-effects regression (nlme::lme or lme4::lmer)", "Normal"),
        ("", ""),
        ("4. Fit the Mortality Model", "Normal"),
        ("   - Track which trees lived vs. died between measurements", "Normal"),
        ("   - Logistic regression: Survival ~ DBH + BA + Relative Height", "Normal"),
        ("", ""),
        ("5. Develop Height-Diameter Curves", "Normal"),
        ("   - Non-linear regression for Curtis-Arney parameters", "Normal"),
        ("   - Fit separately for each species", "Normal"),
        ("", ""),
        ("6. Establish Site Index Curves", "Normal"),
        ("   - Use species-specific growth-and-yield studies", "Normal"),
        ("   - Default SI = regional average productivity", "Normal"),
        ("", ""),
        ("EXAMPLE R CODE FOR MODEL FITTING", "Section"),
        ("", ""),
        ("# Load required libraries", "Code"),
        ("library(nlme)", "Code"),
        ("library(lme4)", "Code"),
        ("library(dplyr)", "Code"),
        ("", ""),
        ("# Crown Ratio Model (cross-sectional)", "Code"),
        ("fit_crown <- nls(cr ~ 10*(b1/(1+b2*ba) + b3*(1-exp(b4*dbh))),", "Code"),
        ("  data = crown_data,", "Code"),
        ("  start = list(b1=5, b2=0.005, b3=3, b4=-0.1),", "Code"),
        ("  subset = species == target_species)", "Code"),
        ("", ""),
        ("# Diameter Growth Model (mixed-effects, repeated measures)", "Code"),
        ("dg_data <- dg_data %>%", "Code"),
        ("  filter(species == target_species) %>%", "Code"),
        ("  mutate(log_dds = log(dds), log_dbh = log(dbh), log_ba = log(ba))", "Code"),
        ("", "Code"),
        ("fit_dg <- lme(log_dds ~ log_dbh + ba + si,", "Code"),
        ("  random = ~ 1|plot,", "Code"),
        ("  data = dg_data)", "Code"),
        ("", ""),
        ("# Mortality Model (logistic regression)", "Code"),
        ("mort_data <- mort_data %>%", "Code"),
        ("  filter(species == target_species) %>%", "Code"),
        ("  mutate(log_dbh = log(dbh), rel_height = height / mean(height))", "Code"),
        ("", "Code"),
        ("fit_mort <- glm(survived ~ log_dbh + ba + rel_height,", "Code"),
        ("  family = binomial(link='logit'),", "Code"),
        ("  data = mort_data)", "Code"),
        ("", ""),
        ("HEIGHT-DIAMETER REGRESSION", "Code"),
        ("fit_htdbh <- nls(height ~ 4.5 + exp(b0 + b1/dbh^b2),", "Code"),
        ("  data = ht_data[ht_data$species == target_species,],", "Code"),
        ("  start = list(b0=-2.5, b1=15, b2=1.0))", "Code"),
        ("", ""),
        ("DATA SOURCE RECOMMENDATIONS", "Section"),
        ("* USDA Forest Service FIA Program", "Normal"),
        ("  - Publicly available plot data with measurements", "Normal"),
        ("  - Multiple measurement occasions for growth analysis", "Normal"),
        ("  - Website: https://www.fia.fs.fed.us/", "Normal"),
        ("", ""),
        ("* Regional Forest Inventory Programs", "Normal"),
        ("  - State forestry agencies", "Normal"),
        ("  - Forest certification programs", "Normal"),
        ("  - University forestry departments", "Normal"),
        ("", ""),
        ("VALIDATION", "Section"),
        ("After calibration, validate your variant by:", "Normal"),
        ("1. Comparing predictions to independent plot data", "Normal"),
        ("2. Checking that projections follow expected trends", "Normal"),
        ("3. Running FVS on regional inventories and reviewing outputs", "Normal"),
        ("4. Performing sensitivity analysis (varying inputs slightly)", "Normal"),
        ("5. Consulting with regional forest ecologists", "Normal"),
        ("", ""),
        ("SUBMITTING YOUR VARIANT", "Section"),
        ("To submit a new variant to the FVS project:", "Normal"),
        ("1. Complete all calibration", "Normal"),
        ("2. Write documentation including:", "Normal"),
        ("   - Species list and FIA mappings", "Normal"),
        ("   - Calibration data sources", "Normal"),
        ("   - Model equations and coefficients", "Normal"),
        ("   - Validation results", "Normal"),
        ("3. Submit to: fvs-feedback@fs.fed.us", "Normal"),
    ]

    for row_num, (text, style_name) in enumerate(instructions, start=1):
        cell = ws.cell(row=row_num, column=1)
        cell.value = text
        if style_name == "Title":
            cell.font = Font(size=16, bold=True)
        elif style_name == "Section":
            cell.font = Font(size=13, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        elif style_name == "Code":
            cell.font = Font(name="Courier New", size=10)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def main():
    """Generate the coefficient template Excel file."""
    output_file = "variant_coefficients_template.xlsx"
    if len(sys.argv) > 1:
        output_file = sys.argv[1]

    wb = Workbook()
    wb.remove(wb.active)  # Remove default blank sheet

    # Create all sheets
    create_species_list_sheet(wb)
    create_crown_coefficients_sheet(wb)
    create_dg_coefficients_sheet(wb)
    create_mortality_coefficients_sheet(wb)
    create_htdbh_coefficients_sheet(wb)
    create_site_index_sheet(wb)
    create_instructions_sheet(wb)

    # Save the workbook
    wb.save(output_file)
    print(f"Created: {output_file}")
    print(f"Sheets created: {len(wb.sheetnames)}")
    print(f"Sheet names: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
